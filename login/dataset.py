import openpyxl
def getRowCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return (sheet.max_row)
def getcolumnCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return (sheet.max_col)
def readdata(file,sheetname,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum,column=columnno).value
def writedata(file,sheetname,rownum,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum,column=columnno).value=data
    workbook.save(file)