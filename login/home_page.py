import time
import pytest
from selenium.webdriver.common.by import By

from login.Base_class import BaseClass


class HomePage(BaseClass):

    def __init__(self,driver):
        self.driver = driver


    def test_login(self):
        import openpyxl
        wb = openpyxl.load_workbook("C:\\Users\\Admin\\Downloads\\login_testcase.xlsx")
        wa = wb.active
        maxrow = wa.max_row
        maxcol = wa.max_column
        for i in range(2, maxrow+1):
            email = wa.cell(i, 1).value
            password = wa.cell(i, 2).value
            print(email, password)


            self.driver.find_element(By.XPATH, "//input[@placeholder='Email']").send_keys(email)
            print(self.driver.find_element(By.XPATH, "//input[@placeholder='Email']").get_attribute("placeholder"))
            self.driver.find_element(By.XPATH, "//input[@placeholder='Password']").send_keys(password)
            self.driver.maximize_window()
            print(self.driver.find_element(By.XPATH, "//input[@placeholder='Password']").get_attribute("placeholder"))
            self.driver.find_element(By.XPATH, "//button[@id='customer-loginbtn']").click()

            time.sleep(10)
            if "Accounts | InsuredMine" in self.driver.title:
                wa.cell(i, 3).value="test passed"
                time.sleep(20)
                self.driver.find_element(By.XPATH,"//a[@class='rounded-bg']/img").click()
                self.driver.find_element(By.XPATH, "//a[@class='rounded-bg']/img").click()
                time.sleep(2)
                self.driver.find_element(By.XPATH, "//div[normalize-space()='Logout']").click()
                self.driver.find_element(By.XPATH, "//button[normalize-space()='Yes, Logout!']").click()
                time.sleep(2)

            else:
                wa.cell(i, 3).value = "test failed"
                self.driver.find_element(By.XPATH, "//input[@placeholder='Email']").clear()
                self.driver.find_element(By.XPATH, "//input[@placeholder='Password']").clear()
        wb.save("result3.xlsx")
    def test_create_account(self):
        try:
            log=self.getLogger()
            log.info("Creating account from account module")
            self.driver.find_element(By.XPATH, "//button[normalize-space()='Add Account']").click()
            log.info("clicking on adding account")
            time.sleep(3)
            print(self.driver.find_element(By.XPATH, "//input[@placeholder='Account Name']").get_attribute("placeholder"))

            self.driver.find_element(By.XPATH, "//input[@placeholder='Account Name']").send_keys("rakeshmirji")
            log.info("entering account name in account name field")

            self.driver.find_element(By.XPATH, "//input[@placeholder='Account Name']").click()
            time.sleep(3)
            self.driver.find_element(By.XPATH,'''//span[normalize-space()='Create New "rakeshmirji"']''').click()
            time.sleep(3)
            try:
                self.driver.find_element(By.XPATH, "//button[@type='button'][normalize-space()='Save']").click()
                log.info("able to click on  save button")
            except Exception as e:
                print(e)
                log.error("Failed to save")

        except Exception as e:
            print(e)